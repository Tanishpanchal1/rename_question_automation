"""
PW Screening — Task 03: Fix the Broken Workflow
================================================
Renames randomly-named images from the ZIP into:
    Q1.png, S1.png, Q2.png, S2.png ... Q45.png, S45.png

Inputs (edit paths below if needed):
    ZIP_PATH  — the ZIP of random-named images
    XLSX_PATH — the Excel metadata file
    OUT_DIR   — folder where renamed files will be saved

Usage:
    pip install openpyxl
    python rename_questions.py
"""

import zipfile
import shutil
import openpyxl
from pathlib import Path

# ── Configure paths ──────────────────────────────────────────
ZIP_PATH  = "practice_zip.zip"
XLSX_PATH = "practice_excel.xlsx"
OUT_DIR   = Path("renamed_questions")
# ─────────────────────────────────────────────────────────────

OUT_DIR.mkdir(exist_ok=True)

# Step 1: Read Excel → build {question_number: filename_id}
wb      = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws      = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

mapping = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    row = list(row) + [None] * (len(headers) - len(row))  # pad short rows
    q_num = row[headers.index("Display Order*")]
    q_img = row[headers.index("Question Image")]
    if q_num and q_img:
        file_id = str(q_img).replace("QUES_ENG_", "").replace(".png", "")
        mapping[int(q_num)] = file_id

print(f"Loaded {len(mapping)} questions from Excel\n")

# Step 2: Extract from ZIP and rename
renamed, errors = [], []

with zipfile.ZipFile(ZIP_PATH, 'r') as zf:
    zip_files = set(zf.namelist())

    for q_num, file_id in sorted(mapping.items()):
        # Question image
        ques_src = f"QUES_ENG_{file_id}.png"
        ques_dst = OUT_DIR / f"Q{q_num}.png"
        if ques_src in zip_files:
            with zf.open(ques_src) as s, open(ques_dst, 'wb') as d:
                shutil.copyfileobj(s, d)
            renamed.append(f"Q{q_num}.png  ←  {ques_src}")
        else:
            errors.append(f"MISSING QUES: {ques_src}")

        # Solution image (shares the same ID, just SOLU_ prefix)
        solu_src = f"SOLU_ENG_{file_id}.png"
        solu_dst = OUT_DIR / f"S{q_num}.png"
        if solu_src in zip_files:
            with zf.open(solu_src) as s, open(solu_dst, 'wb') as d:
                shutil.copyfileobj(s, d)
            renamed.append(f"S{q_num}.png  ←  {solu_src}")
        else:
            errors.append(f"MISSING SOLU: {solu_src}")

# Step 3: Report
print(f"{'='*55}")
for r in renamed:
    print(f"  ✓  {r}")

print(f"\n{'='*55}")
print(f"  Total renamed : {len(renamed)} files")
print(f"  Errors        : {len(errors)}")

if errors:
    print("\nMissing files:")
    for e in errors:
        print(f"  ✗  {e}")
else:
    print("  Zero errors — perfect match ✅")

print(f"\nOutput folder: {OUT_DIR.resolve()}")
