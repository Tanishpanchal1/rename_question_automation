import sys, os
from pathlib import Path
import openpyxl, json
excel_path = Path(r'D:\Task-3\Practice Test - 08_Eng._Physics_Lakshya NEET Hindi Test Series (2026)_21-12-2025_Excel (1).xlsx')
wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print('Headers:', headers)
rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
    rows.append(row)
print('First rows:', rows)
