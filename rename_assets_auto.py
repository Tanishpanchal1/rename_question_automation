#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
rename_assets_auto.py
--------------------
Automatically renames the question and solution screenshots from the images folder
based on the "Display Order*" defined in the corresponding Excel file.

Steps:
1. Load Excel worksheet and locate "Display Order*" and "QBG Question id" columns.
2. Build a mapping of display orders to question hashes.
3. For each pair, locate the QUES_ENG_<hash> and SOLU_ENG_<hash> images.
4. Copy them to the output "renamed_assets" folder as Q<order>.<ext> and S<order>.<ext>.
"""

import os
import shutil
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------
BASE_DIR = Path(r"d:\Task-3")
ASSETS_ROOT = BASE_DIR / "Practice Test - 08_Eng._Physics_Lakshya NEET Hindi Test Series (2026)_21-12-2025_Image (1)"
EXCEL_PATH = BASE_DIR / "Practice Test - 08_Eng._Physics_Lakshya NEET Hindi Test Series (2026)_21-12-2025_Excel (1).xlsx"
OUTPUT_DIR = BASE_DIR / "renamed_assets"

QUESTION_PREFIX = "QUES_ENG_"
SOLUTION_PREFIX = "SOLU_ENG_"

def find_image(directory: Path, prefix: str, qbg_id: str) -> Path:
    """Finds an image with the given prefix and ID, dynamically matching any extension."""
    pattern = f"{prefix}{qbg_id}.*"
    matches = list(directory.glob(pattern))
    if not matches:
        return None
    return matches[0]

def main() -> int:
    # -------------------------------------------------------------------
    # Validation of paths
    # -------------------------------------------------------------------
    if not ASSETS_ROOT.is_dir():
        print(f"[ERROR] Expected images folder not found at: {ASSETS_ROOT}")
        return 1
    if not EXCEL_PATH.is_file():
        print(f"[ERROR] Expected Excel sheet not found at: {EXCEL_PATH}")
        return 1

    print(f"[INFO] Reading Excel file: {EXCEL_PATH.name}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        print(f"[ERROR] Error loading Excel workbook: {e}")
        return 1

    # Get header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        display_order_idx = headers.index("Display Order*")
        qbg_id_idx = headers.index("QBG Question id")
    except ValueError as e:
        print(f"[ERROR] Required columns ('Display Order*' or 'QBG Question id') not found in headers: {headers}")
        return 1

    # -------------------------------------------------------------------
    # Build mapping from Excel sheet
    # -------------------------------------------------------------------
    print("[INFO] Parsing question display order mappings...")
    mappings = []
    seen_orders = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # Skip completely empty rows
        
        order_val = row[display_order_idx]
        qbg_id = row[qbg_id_idx]
        
        if order_val is None or qbg_id is None:
            print(f"[WARN] Row {row_num}: Skipping due to missing Display Order ({order_val}) or QBG Question id ({qbg_id})")
            continue

        try:
            # Convert float/string order to clean integer
            order_int = int(float(order_val))
        except (ValueError, TypeError):
            print(f"[WARN] Row {row_num}: Skipping due to invalid display order format: {order_val}")
            continue

        if order_int in seen_orders:
            print(f"[WARN] Row {row_num}: Duplicate Display Order detected ({order_int}) for question ID {qbg_id}. Skipping.")
            continue
            
        seen_orders.add(order_int)
        mappings.append((order_int, qbg_id))

    # Sort mappings by display order
    mappings.sort(key=lambda x: x[0])
    print(f"[INFO] Found {len(mappings)} valid mappings in Excel sheet.")

    # -------------------------------------------------------------------
    # Recreate clean output directory (Windows-safe method)
    # -------------------------------------------------------------------
    if OUTPUT_DIR.exists():
        print(f"[INFO] Clearing existing output folder: {OUTPUT_DIR.name}")
        for item in OUTPUT_DIR.iterdir():
            try:
                if item.is_file() or item.is_symlink():
                    item.unlink()
                elif item.is_dir():
                    shutil.rmtree(item)
            except Exception as e:
                print(f"[WARN] Could not clean file {item.name}: {e}")
    else:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # -------------------------------------------------------------------
    # Rename and Copy files
    # -------------------------------------------------------------------
    print("\n[INFO] Copying and renaming files...")
    success_count = 0
    missing_questions = []
    missing_solutions = []

    for order, qbg_id in mappings:
        q_src = find_image(ASSETS_ROOT, QUESTION_PREFIX, qbg_id)
        s_src = find_image(ASSETS_ROOT, SOLUTION_PREFIX, qbg_id)

        if not q_src:
            missing_questions.append((order, qbg_id))
            print(f"[ERROR] Question image missing for Order {order} (ID: {qbg_id})")
            continue
        
        if not s_src:
            missing_solutions.append((order, qbg_id))
            print(f"[ERROR] Solution image missing for Order {order} (ID: {qbg_id})")
            continue

        # Determine target destinations
        q_dst = OUTPUT_DIR / f"Q{order}{q_src.suffix.lower()}"
        s_dst = OUTPUT_DIR / f"S{order}{s_src.suffix.lower()}"

        try:
            shutil.copy2(q_src, q_dst)
            shutil.copy2(s_src, s_dst)
            print(f"[OK] Row mapped: {q_src.name} -> {q_dst.name}  |  {s_src.name} -> {s_dst.name}")
            success_count += 1
        except Exception as e:
            print(f"[ERROR] Failed to copy Order {order} images: {e}")

    # -------------------------------------------------------------------
    # Summary
    # -------------------------------------------------------------------
    print("\n" + "=" * 50)
    print("                       SUMMARY")
    print("=" * 50)
    print(f"Total mapped questions processed: {len(mappings)}")
    print(f"Successfully renamed pairs:       {success_count}")
    
    if missing_questions:
        print(f"Missing Question files count:     {len(missing_questions)}")
    if missing_solutions:
        print(f"Missing Solution files count:     {len(missing_solutions)}")

    if success_count == len(mappings) and not missing_questions and not missing_solutions:
        print("\n[SUCCESS] All assets renamed successfully without any missing files.")
        print(f"Renamed files are located in: {OUTPUT_DIR.resolve()}\n")
        return 0
    else:
        print("\n[WARNING] Done with warnings/errors. Please check the logs above.\n")
        return 1

if __name__ == "__main__":
    raise SystemExit(main())
